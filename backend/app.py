from flask import Flask, request, jsonify, make_response
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date ,timedelta
import os
import json
import logging
from logging.handlers import RotatingFileHandler
from models import db, Applicant, Job, Employee, Interview, PhoneCall
import re
import traceback

# 変更箇所
import os
from dotenv import load_dotenv



# アプリケーションの初期化
app = Flask(__name__)

# .envファイルから環境変数を読み込む
load_dotenv()

# 環境設定
ENV = os.environ.get('FLASK_ENV', 'development')
DEBUG = ENV == 'development'

# Google Maps APIキーを環境変数から取得
app.config['GOOGLE_MAPS_API_KEY'] = os.environ.get('GOOGLE_MAPS_API_KEY')


# アプリケーションの初期化
app = Flask(__name__)

# 環境設定
ENV = os.environ.get('FLASK_ENV', 'development')
DEBUG = ENV == 'development'

# ロギング設定
if not os.path.exists('logs'):
    os.mkdir('logs')

file_handler = RotatingFileHandler('logs/app.log', maxBytes=10240, backupCount=10)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
))
file_handler.setLevel(logging.INFO)
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO if DEBUG else logging.WARNING)
app.logger.info('アプリケーション起動')

# リクエストロギング用のデバッグミドルウェア
@app.before_request
def log_request_info():
    app.logger.info(f'リクエスト: {request.method} {request.path}')

# CORS設定
if DEBUG:
    # 開発環境 - より寛容な設定
    CORS(app, resources={r"/api/*": {
        "origins": ["http://localhost:3000", "http://127.0.0.1:3000"],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization", "Accept"],
        "supports_credentials": False
    }})
else:
    # 本番環境 - 制限された設定
    CORS(app, resources={r"/api/*": {
        "origins": [os.environ.get('ALLOWED_ORIGIN', 'https://your-production-domain.com')],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization", "Accept"],
        "supports_credentials": False
    }})

# データベース設定
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///recruiting.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

# セキュリティ設定
app.config['JSON_SORT_KEYS'] = False
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = DEBUG

# クロスドメインアクセス用のデコレータ
def add_cors_headers(response):
    allowed_origins = ["http://localhost:3000", "http://127.0.0.1:3000"]
    if ENV == 'production':
        allowed_origins = [os.environ.get('ALLOWED_ORIGIN', 'https://your-production-domain.com')]
    
    origin = request.headers.get('Origin')
    if origin and origin in allowed_origins:
        response.headers.add('Access-Control-Allow-Origin', origin)
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization,Accept')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# OPTIONSリクエストのハンドラー
@app.route('/api/<path:path>', methods=['OPTIONS'])
def options_handler(path):
    app.logger.info(f'OPTIONS リクエスト処理: /api/{path}')
    response = make_response()
    return add_cors_headers(response)

# データベーステーブルを作成する関数
def create_tables():
    db.reflect()
    db.create_all()

# アプリケーションコンテキスト内でデータベースを作成
with app.app_context():
    create_tables()

# 文字列を日付オブジェクトに変換する関数
def parse_date(date_str, format='%Y-%m-%d'):
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, format).date()
    except ValueError:
        app.logger.warning(f"Invalid date format: {date_str}, expected format: {format}")
        return None

# ホームページ
@app.route("/")
def hello_world():
    response = make_response("<p>人材紹介アプリケーションAPI</p>")
    return add_cors_headers(response)

# データベースを初期化するエンドポイント
@app.route('/api/init-db', methods=['POST'])
def init_db():
    db.drop_all()
    db.create_all()
    response = make_response(jsonify({'message': 'Database initialized successfully'}), 201)
    return add_cors_headers(response)


# 求職者一覧取得API（エラーハンドリング強化版）
@app.route('/api/applicants', methods=['GET'])
def get_applicants():
    try:
        # データベースから求職者情報を取得（ページネーション）
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)

        app.logger.info(f"求職者一覧取得リクエスト: ページ={page}, 件数={per_page}")

        # クエリの実行
        try:
            # SQLAlchemyのバージョンに応じたページネーション
            try:
                applicants = Applicant.query.paginate(page=page, per_page=per_page, error_out=False)
            except TypeError:
                # 古いSQLAlchemyのバージョン用
                applicants = Applicant.query.paginate(page, per_page, False)

            app.logger.info(f"求職者データ取得成功: {len(applicants.items)}件")

            # JSONに変換
            result = []
            for applicant in applicants.items:
                try:
                    # 通常の変換を試行
                    applicant_dict = applicant.to_dict()
                    
                    # 代替フィールドからの担当社員情報抽出
                    if hasattr(applicant, 'important_point_details') and applicant.important_point_details:
                        if '【担当社員ID】' in applicant.important_point_details:
                            match = re.search(r'【担当社員ID】(\d+)', applicant.important_point_details)
                            if match:
                                applicant_dict['assigned_employee_id'] = int(match.group(1))
                    
                    # バックアップフィールドからの抽出
                    if 'assigned_employee_id' not in applicant_dict and hasattr(applicant, 'most_important_point') and applicant.most_important_point:
                        if '担当社員ID:' in applicant.most_important_point:
                            match = re.search(r'担当社員ID:(\d+)', applicant.most_important_point)
                            if match:
                                applicant_dict['assigned_employee_id'] = int(match.group(1))
                    
                    result.append(applicant_dict)
                except Exception as conversion_error:
                    app.logger.error(f"求職者データ変換エラー ID={applicant.id}: {str(conversion_error)}")
                    # 最小限の情報のみ返す
                    result.append({
                        'id': applicant.id,
                        'name': getattr(applicant, 'name', '不明'),
                        'error': f'データ変換エラー: {str(conversion_error)}'
                    })

            response = make_response(jsonify(result))
            return add_cors_headers(response)

        except Exception as query_error:
            app.logger.error(f"求職者クエリエラー: {str(query_error)}\n{traceback.format_exc()}")
            # 代替として、より少ないデータを返す
            limited_applicants = Applicant.query.limit(10).all()
            basic_data = []
            
            for applicant in limited_applicants:
                try:
                    app_dict = {
                        'id': applicant.id,
                        'name': getattr(applicant, 'name', '不明'),
                        'note': '技術的な問題により、完全なデータを取得できません'
                    }
                    
                    # 電話番号があれば追加
                    if hasattr(applicant, 'phone_number') and applicant.phone_number:
                        app_dict['phone_number'] = applicant.phone_number
                    
                    # メールアドレスがあれば追加
                    if hasattr(applicant, 'email') and applicant.email:
                        app_dict['email'] = applicant.email
                        
                    basic_data.append(app_dict)
                except Exception as item_error:
                    app.logger.error(f"求職者データ基本情報エラー ID={applicant.id}: {str(item_error)}")
                    basic_data.append({
                        'id': applicant.id,
                        'name': '不明',
                        'error': str(item_error)
                    })

            response = make_response(jsonify(basic_data))
            return add_cors_headers(response)

    except Exception as e:
        app.logger.exception(f"求職者一覧取得エラー: {str(e)}\n{traceback.format_exc()}")
        response = make_response(jsonify({
            'error': str(e),
            'message': '求職者データの取得中にサーバーエラーが発生しました'
        }), 500)
        return add_cors_headers(response)


# 求職者追加API
@app.route('/api/applicants', methods=['POST'])
def add_applicant():
    try:
        data = request.json
        app.logger.info(f"求職者追加リクエスト: {data}")
        
        # 日付文字列をDateオブジェクトに変換
        birthdate = None
        if 'birthdate' in data and data['birthdate']:
            birthdate = parse_date(data['birthdate'])
            if birthdate is None:
                return jsonify({'error': 'Invalid date format for birthdate. Use YYYY-MM-DD'}), 400
        
        # 新しい求職者を作成
        applicant = Applicant(
            name=data.get('name'),
            address=data.get('address'),
            desired_occupation=data.get('desired_occupation'),
            desired_location=data.get('desired_location'),
            birthdate=birthdate,
            email=data.get('email'),
            phone_number=data.get('phone_number'),
            gender=data.get('gender'),
            nationality=data.get('nationality'),
            employment_status=data.get('employment_status'),
            available_date=data.get('available_date'),
            employment_period=data.get('employment_period'),
            medical_history=data.get('medical_history'),
            disability_certificate=data.get('disability_certificate'),
            tattoo=data.get('tattoo'),
            tattoo_details=data.get('tattoo_details'),
            criminal_record=data.get('criminal_record'),
            clothing_size=data.get('clothing_size'),
            commute_or_dormitory=data.get('commute_or_dormitory'),
            commute_method=data.get('commute_method'),
            commute_area=data.get('commute_area'),
            factory_experience=data.get('factory_experience'),
            experience_details=data.get('experience_details'),
            desired_working_hours=data.get('desired_working_hours'),
            recent_applications=data.get('recent_applications'),
            most_important_point=data.get('most_important_point'),
            important_point_details=data.get('important_point_details'),
            desired_salary=data.get('desired_salary'),
            height=data.get('height'),
            weight=data.get('weight')
        )
        
        db.session.add(applicant)
        db.session.commit()
        
        app.logger.info(f"求職者追加成功 ID: {applicant.id}")
        response = make_response(jsonify({
            'message': 'Applicant added successfully',
            'applicant': applicant.to_dict()
        }), 201)
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# 求職者更新API
@app.route('/api/applicants/<int:applicant_id>', methods=['PUT'])
def update_applicant(applicant_id):
    try:
        applicant = Applicant.query.get(applicant_id)
        if not applicant:
            response = make_response(jsonify({'error': 'Applicant not found'}), 404)
            return add_cors_headers(response)
        
        data = request.json
        app.logger.info(f"求職者更新リクエスト ID: {applicant_id}, データ: {data}")
        
        # 日付文字列をDateオブジェクトに変換
        if 'birthdate' in data and data['birthdate']:
            birthdate = parse_date(data['birthdate'])
            if birthdate is None:
                response = make_response(jsonify({'error': 'Invalid date format for birthdate. Use YYYY-MM-DD'}), 400)
                return add_cors_headers(response)
            applicant.birthdate = birthdate
        
        # 各フィールドを更新
        for field in [
            'name', 'address', 'desired_occupation', 'desired_location', 'email', 'phone_number', 'gender', 'nationality',
            'employment_status', 'available_date', 'employment_period', 'medical_history',
            'disability_certificate', 'tattoo', 'tattoo_details', 'criminal_record', 'clothing_size',
            'commute_or_dormitory', 'commute_method', 'commute_area', 'factory_experience',
            'experience_details', 'desired_working_hours', 'recent_applications', 'most_important_point',
            'important_point_details', 'desired_salary', 'height', 'weight'
        ]:
            if field in data:
                setattr(applicant, field, data[field])
        
        db.session.commit()
        app.logger.info(f"求職者更新成功 ID: {applicant_id}")
        response = make_response(jsonify({
            'message': 'Applicant updated successfully',
            'applicant': applicant.to_dict()
        }))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# 求職者削除API
@app.route('/api/applicants/<int:applicant_id>', methods=['DELETE'])
def delete_applicant(applicant_id):
    try:
        applicant = Applicant.query.get(applicant_id)
        if not applicant:
            response = make_response(jsonify({'error': 'Applicant not found'}), 404)
            return add_cors_headers(response)
        
        db.session.delete(applicant)
        db.session.commit()
        app.logger.info(f"求職者削除成功 ID: {applicant_id}")
        response = make_response(jsonify({'message': 'Applicant deleted successfully'}))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# 求職者進捗更新API
@app.route('/api/applicants/<int:applicant_id>/progress', methods=['PUT'])
def update_applicant_progress(applicant_id):
    try:
        applicant = Applicant.query.get(applicant_id)
        if not applicant:
            response = make_response(jsonify({'error': 'Applicant not found'}), 404)
            return add_cors_headers(response)
        
        data = request.json
        app.logger.info(f"求職者進捗更新リクエスト ID: {applicant_id}, データ: {data}")
        
        # 変更前の値をログに記録（デバッグ用）
        old_values = {}
        for field in ['application_date', 'call_date', 'connection_date', 'proposal_date', 'document_sent_date', 'document_passed_date', 'interview_date', 'offer_date', 'hire_date', 'payment_date']:
            old_values[field] = getattr(applicant, field)
        
        # 進捗データの更新
        for field in ['application_date', 'call_date', 'connection_date', 'proposal_date', 'document_sent_date', 'document_passed_date', 'interview_date', 'offer_date', 'hire_date', 'payment_date']:
            if field in data:
                # 日付データの処理
                if data[field]:
                    try:
                        # 日付文字列から日付オブジェクトへの変換
                        date_value = datetime.strptime(data[field], '%Y-%m-%d').date()
                        setattr(applicant, field, date_value)
                        app.logger.debug(f"フィールド {field} を {data[field]} から {date_value} に更新")
                    except ValueError as e:
                        app.logger.error(f"日付変換エラー: {e}, value: {data[field]}")
                        return jsonify({'error': f'Invalid date format for {field}. Use YYYY-MM-DD'}), 400
                else:
                    setattr(applicant, field, None)
                    app.logger.debug(f"フィールド {field} を None に設定")
        
        # 変更が実際に行われたか確認
        db.session.flush()
        
        # 変更後の値をログに記録（デバッグ用）
        new_values = {}
        for field in ['application_date', 'call_date', 'connection_date', 'proposal_date', 'document_sent_date', 'document_passed_date', 'interview_date', 'offer_date', 'hire_date', 'payment_date']:
            new_values[field] = getattr(applicant, field)
            if old_values[field] != new_values[field]:
                app.logger.info(f"フィールド {field} の値が変更されました: {old_values[field]} -> {new_values[field]}")
        
        # トランザクションをコミット
        db.session.commit()
        
        app.logger.info(f"求職者進捗更新成功 ID: {applicant_id}")
        
        # 更新された求職者データを返す
        updated_applicant_dict = applicant.to_dict()
        app.logger.debug(f"更新後の求職者データ: {updated_applicant_dict}")
        
        response = make_response(jsonify({
            'message': 'Progress updated successfully',
            'applicant': updated_applicant_dict
        }))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)
    
# 個別の求職者詳細を取得するAPIエンドポイント
@app.route('/api/applicants/<int:applicant_id>', methods=['GET'])
def get_applicant(applicant_id):
    try:
        applicant = Applicant.query.get(applicant_id)
        if not applicant:
            response = make_response(jsonify({'error': 'Applicant not found'}), 404)
            return add_cors_headers(response)

        app.logger.info(f"求職者詳細取得成功 ID: {applicant_id}")
        
        # 通常のto_dict()結果を取得
        applicant_data = applicant.to_dict()
        
        # ここでassigned_employee_idを明示的に追加
        applicant_data['assigned_employee_id'] = getattr(applicant, 'assigned_employee_id', None)
        applicant_data['employee_id'] = getattr(applicant, 'assigned_employee_id', None)  # 互換性のため両方設定
        
        response = make_response(jsonify(applicant_data))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)



# 求職者紹介料更新API
@app.route('/api/applicants/<int:applicant_id>/referral-fee', methods=['PUT'])
def update_applicant_referral_fee(applicant_id):
    try:
        applicant = Applicant.query.get(applicant_id)
        if not applicant:
            response = make_response(jsonify({'error': 'Applicant not found'}), 404)
            return add_cors_headers(response)
        
        data = request.json
        app.logger.info(f"求職者紹介料更新リクエスト ID: {applicant_id}, データ: {data}")
        
        if 'referral_fee' not in data:
            response = make_response(jsonify({'error': 'Referral fee is required'}), 400)
            return add_cors_headers(response)
        
        # 数値として保存
        try:
            referral_fee = int(re.sub(r'[^\d]', '', str(data['referral_fee'])))
            applicant.referral_fee = referral_fee
        except ValueError:
            response = make_response(jsonify({'error': 'Invalid referral fee value'}), 400)
            return add_cors_headers(response)
        
        db.session.commit()
        app.logger.info(f"求職者紹介料更新成功 ID: {applicant_id}, 金額: {applicant.referral_fee}")
        
        response = make_response(jsonify({
            'message': 'Referral fee updated successfully',
            'applicant': applicant.to_dict()
        }))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# 求人情報関連のAPI
@app.route('/api/upload-job-data', methods=['POST'])
def upload_job_data():
    if 'file' not in request.files:
        response = make_response(jsonify({'error': 'No file part'}), 400)
        return add_cors_headers(response)

    file = request.files['file']
    if file.filename == '':
        response = make_response(jsonify({'error': 'No selected file'}), 400)
        return add_cors_headers(response)

    sheet_type = request.form.get('sheetType', 'cnt')  # デフォルトは'cnt'
    if file:
        try:
            import openpyxl
            app.logger.info(f"Excelファイルアップロード: {file.filename}, シートタイプ: {sheet_type}")

            # openpyxlを使用してファイルを読み込む
            workbook = openpyxl.load_workbook(file)
            
            # シートタイプに基づいて適切なシートを選択
            sheet_name = "新案件共有データ(CNT)最新" if sheet_type == 'cnt' else "新案件共有データ(AIM)最新"
            
            # 指定されたシートが存在するか確認
            if sheet_name not in workbook.sheetnames:
                # シートが見つからない場合はアクティブなシートを使用
                worksheet = workbook.active
                app.logger.warning(f"指定されたシート '{sheet_name}' が見つかりません。アクティブなシートを使用します。")
                response = make_response(jsonify({
                    'warning': f'指定されたシート "{sheet_name}" が見つかりませんでした。アクティブなシートを使用します。'
                }), 200)
                return add_cors_headers(response)
            else:
                worksheet = workbook[sheet_name]

            # ヘッダー行を取得（6行目がヘッダー）
            headers_raw = [cell.value for cell in worksheet[6]]
            
            # ヘッダーの標準化（改行や余分なスペースの除去）
            headers = []
            for header in headers_raw:
                if header is not None:
                    # 改行、余分な空白を削除して標準化する
                    cleaned_header = re.sub(r'\s+', ' ', str(header)).strip()
                    headers.append(cleaned_header)
                else:
                    headers.append(None)
            
            app.logger.debug(f"標準化されたヘッダー行: {headers}")

            # CNTシートとAIMシートのマッピングテーブル
            field_mappings = {
                'cnt': {
                    'job_url': 'ジョブパルＵＲＬ',
                    'job_number': 'お仕事№',
                    'cf_fc': 'cf / fc / 事業所',
                    'company': '企業名 / 工場名',
                    'prefecture': '所在地 （都道府県）',
                    'city': '所在地 （市区町村以降）',
                    'salary': '総支給額',
                    'fee': 'フィー',
                    'age_limit': '年齢上限1',
                    'min_age': '年齢下限1',
                    'description': '業務内容詳細',
                    'requirements': '配属可能条件に関する備考1',
                    'benefits': '特別待遇',
                    'working_hours': '勤務時間に関する備考',
                    'employment_type': '雇用形態',
                    'holidays': '休日',
                    'dormitory': '入寮可否',
                    'housing_cost': '社宅費負担',
                    'housing_allowance': '社宅費補助額',
                    'work_style': '勤務形態',
                    'annual_holidays': '年間休日',
                    'gender': '性別1',
                    'work_experience': '業務経験1',
                    'occupation_experience': '職種経験1',
                    'japanese_required': '外国人受け入れ',
                    'commute_method': '可能通勤手段',
                    'nearest_station': '最寄り駅（駅名）',
                    'salary_type': '給与形態',
                    'hourly_wage': '時給',
                    'shift': 'シフト',
                    'products': '生産品目',
                    'occupation_major_category': '職種①大分類',
                    'occupation_minor_category': '職種①小分類',
                    'advantages': 'メリット（訴求ポイント）',
                    'smoking_measures': '受動喫煙防止対策'
                },
                'aim': {
                    'job_url': 'ジョブパルＵＲＬ',
                    'job_number': 'お仕事№',
                    'cf_fc': 'cf / fc / 事業所',
                    'company': '企業名 / 工場名',
                    'prefecture': '所在地 （都道府県）',
                    'city': '所在地 （市区町村以降）',
                    'salary': '総支給額',
                    'fee': 'Fee',
                    'age_limit': '年齢上限1',
                    'min_age': '年齢下限1',
                    'description': '業務内容詳細',
                    'requirements': '配属可能条件に関する備考1',
                    'benefits': '特別待遇',
                    'working_hours': '勤務時間に関する備考',
                    'employment_type': '雇用形態',
                    'holidays': '休日',
                    'dormitory': '入寮可否',
                    'housing_cost': '社宅費負担',
                    'housing_allowance': '社宅費補助額',
                    'work_style': '勤務形態',
                    'annual_holidays': '年間休日',
                    'gender': '性別1',
                    'work_experience': '業務経験1',
                    'occupation_experience': '職種経験1',
                    'japanese_required': '外国人受け入れ',
                    'commute_method': '可能通勤手段',
                    'nearest_station': '最寄り駅（駅名）',
                    'salary_type': '給与形態',
                    'hourly_wage': '時給',
                    'shift': 'シフト',
                    'products': '生産品目',
                    'occupation_major_category': '職種大分類',
                    'occupation_minor_category': '職種小分類',
                    'advantages': 'メリット（訴求ポイント）',
                    'smoking_measures': '受動喫煙防止対策'
                }
            }
            
            # 実際のヘッダーに基づいてマッピングを調整
            mapping = field_mappings['cnt'] if sheet_type == 'cnt' else field_mappings['aim']
            
            # 標準化されたヘッダーに基づいてマッピングキーを更新
            for model_field, excel_column in list(mapping.items()):
                # 完全一致検索
                if excel_column in headers:
                    continue
                
                # ヘッダーに存在する類似の列名を探す
                for header in headers:
                    if header and (
                        (model_field == 'prefecture' and ('都道府県' in header or '県' in header)) or
                        (model_field == 'city' and ('市区町村' in header or '市' in header)) or
                        (excel_column.replace(' ', '') in header.replace(' ', ''))
                    ):
                        app.logger.info(f"列名マッピングを更新: {model_field} -> {header}")
                        mapping[model_field] = header
                        break
            
            # マッピングのデバッグ出力
            app.logger.info(f"調整後のマッピング: {mapping}")
            
            job_count = 0
            
            # データ行を処理（7行目からデータ開始）
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=7, values_only=True), 7):
                try:
                    # 辞書形式に変換
                    row_data = dict(zip(headers, row))
                    
                    # 必須フィールドのチェック
                    company = row_data.get(mapping['company'], '')
                    if isinstance(company, tuple):
                        company = company[0] if company else ''
                    
                    # 空のデータ行はスキップ
                    if not company:
                        continue
                    
                    # タイトルが存在しない場合は会社名を使用
                    title = company
                    
                    # 所在地データの標準化
                    prefecture_raw = row_data.get(mapping['prefecture'])
                    prefecture = ''
                    if prefecture_raw is not None:
                        if isinstance(prefecture_raw, (str, int, float)):
                            prefecture = str(prefecture_raw).strip()
                        elif isinstance(prefecture_raw, tuple) and len(prefecture_raw) > 0:
                            prefecture = str(prefecture_raw[0]).strip()
                    
                    city_raw = row_data.get(mapping['city'])
                    city = ''
                    if city_raw is not None:
                        if isinstance(city_raw, (str, int, float)):
                            city = str(city_raw).strip()
                        elif isinstance(city_raw, tuple) and len(city_raw) > 0:
                            city = str(city_raw[0]).strip()
                    
                    # 新しいJobオブジェクトを作成
                    job = Job(
                        title=title,
                        company=company,
                        prefecture=prefecture,
                        city=city
                    )
                    
                    # その他のフィールドを設定
                    for model_field, excel_column in mapping.items():
                        # すでに設定済みのフィールドはスキップ
                        if model_field in ['title', 'company', 'prefecture', 'city']:
                            continue
                        
                        # Excelの列が存在する場合のみ処理
                        if excel_column in headers and excel_column in row_data:
                            value = row_data.get(excel_column)
                            
                            # 特別な変換が必要なフィールドの処理
                            if model_field == 'dormitory':
                                # '可'または'有'の場合はTrue、それ以外はFalse
                                if isinstance(value, str):
                                    value = value.strip() in ['可', '有', '◯', '○', '〇', 'あり']
                                else:
                                    value = False
                            elif model_field == 'japanese_required':
                                # '不可'または'無'の場合はTrue（外国人受け入れなし = 日本語必須）
                                if isinstance(value, str):
                                    value = value.strip() in ['不可', '無', '×', '不可能']
                                else:
                                    value = False
                            elif model_field in ['hourly_wage', 'min_age', 'age_limit'] and value:
                                # 数値文字列を整数に変換
                                try:
                                    if isinstance(value, (int, float)):
                                        value = int(value)
                                    else:
                                        # 数字以外の文字を削除して整数に変換
                                        cleaned_value = ''.join(filter(str.isdigit, str(value)))
                                        value = int(cleaned_value) if cleaned_value else None
                                except:
                                    value = None
                            
                            # モデルのフィールドに値を設定（Noneでない場合のみ）
                            if value is not None:
                                setattr(job, model_field, value)
                    
                    # デバッグ用：作成したオブジェクトの内容をログ出力
                    app.logger.debug(f"求人作成: {job.title}, {job.company}, 勤務地: {job.prefecture} {job.city}")
                    
                    db.session.add(job)
                    job_count += 1
                    
                    # 定期的にコミット（大量データの場合のメモリ対策）
                    if job_count % 100 == 0:
                        db.session.commit()
                        app.logger.info(f"{job_count}件の求人情報を処理しました")
                
                except Exception as row_error:
                    app.logger.error(f"行 {row_idx} の処理中にエラー: {str(row_error)}")
                    # 個別の行のエラーは無視して次に進む
                    continue
            
            # 残りのデータをコミット
            db.session.commit()
            
            app.logger.info(f"求人情報インポート成功: {job_count}件")
            
            response = make_response(jsonify({
                'message': f'ファイルがアップロードされ、シート "{sheet_name}" から{job_count}件の求人情報が正常に処理されました',
                'job_count': job_count
            }), 200)
            return add_cors_headers(response)
            
        except Exception as e:
            db.session.rollback()
            traceback_str = traceback.format_exc()
            app.logger.error(f"求人アップロードエラー: {str(e)}\n{traceback_str}")
            response = make_response(jsonify({
                'error': str(e),
                'traceback': traceback_str
            }), 500)
            return add_cors_headers(response)
    
    response = make_response(jsonify({'error': 'File processing failed'}), 500)
    return add_cors_headers(response)

@app.route('/api/jobs', methods=['GET'])
def get_jobs():
    try:
        jobs = Job.query.all()
        app.logger.debug(f"求人情報件数: {len(jobs)}")
        response = make_response(jsonify([job.to_dict() for job in jobs]))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/jobs', methods=['POST'])
def add_job():
    try:
        data = request.json
        app.logger.info(f"求人情報追加リクエスト: {data}")
        
        # 必須フィールドの検証
        if not data.get('title') or not data.get('company'):
            response = make_response(jsonify({'error': 'Title and company are required'}), 400)
            return add_cors_headers(response)
        
        # 新しい求人情報を作成
        job = Job(
            title=data.get('title'),
            company=data.get('company'),
            job_url=data.get('job_url'),
            job_number=data.get('job_number'),
            cf_fc=data.get('cf_fc'),
            prefecture=data.get('prefecture'),
            city=data.get('city'),
            salary=data.get('salary'),
            fee=data.get('fee'),
            age_limit=data.get('age_limit'),
            description=data.get('description'),
            requirements=data.get('requirements'),
            benefits=data.get('benefits'),
            working_hours=data.get('working_hours'),
            employment_type=data.get('employment_type'),
            holidays=data.get('holidays'),
            dormitory=data.get('dormitory'),
            work_style=data.get('work_style'),
            annual_holidays=data.get('annual_holidays'),
            gender=data.get('gender'),
            commute_method=data.get('commute_method'),
            nearest_station=data.get('nearest_station'),
            salary_type=data.get('salary_type'),
            shift=data.get('shift'),
            products=data.get('products'),
            occupation_major_category=data.get('occupation_major_category'),
            occupation_minor_category=data.get('occupation_minor_category'),
            advantages=data.get('advantages'),
            smoking_measures=data.get('smoking_measures')
        )
        
        db.session.add(job)
        db.session.commit()
        app.logger.info(f"求人情報追加成功 ID: {job.id}")
        
        response = make_response(jsonify({
            'message': 'Job added successfully',
            'job': job.to_dict()
        }), 201)
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/jobs/<int:job_id>', methods=['GET'])
def get_job(job_id):
    try:
        app.logger.info(f"求人詳細取得リクエスト: ID={job_id}")
        job = Job.query.get(job_id)
        if job:
            app.logger.info(f"求人詳細取得成功: ID={job_id}, 企業={job.company}")
            response = make_response(jsonify(job.to_dict()))
            return add_cors_headers(response)
        
        app.logger.warning(f"求人詳細取得失敗: ID={job_id} - 該当データなし")
        response = make_response(jsonify({'error': 'Job not found'}), 404)
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(f"求人詳細取得エラー: ID={job_id}, エラー={str(e)}")
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/jobs/delete-all', methods=['DELETE'])
def delete_all_jobs():
    try:
        count = db.session.query(Job).delete()
        db.session.commit()
        app.logger.warning(f"全求人情報削除: {count}件")
        response = make_response(jsonify({'message': f'All jobs deleted successfully ({count} records)'}))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# 社員関連のAPI
@app.route('/api/employees', methods=['GET'])
def get_employees():
    try:
        employees = Employee.query.all()
        response = make_response(jsonify([employee.to_dict() for employee in employees]))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/employees/<int:employee_id>', methods=['GET'])
def get_employee(employee_id):
    try:
        employee = Employee.query.get(employee_id)
        if employee:
            response = make_response(jsonify(employee.to_dict()))
            return add_cors_headers(response)
        response = make_response(jsonify({'error': 'Employee not found'}), 404)
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/employees', methods=['POST'])
def add_employee():
    try:
        data = request.json
        app.logger.info(f"社員追加リクエスト: {data}")
        
        # 必須フィールドの検証
        if not data.get('name'):
            response = make_response(jsonify({'error': 'Name is required'}), 400)
            return add_cors_headers(response)
        
        # 日付文字列をDateオブジェクトに変換
        hire_date = None
        if 'hire_date' in data and data['hire_date']:
            hire_date = parse_date(data['hire_date'])
            if hire_date is None:
                response = make_response(jsonify({'error': 'Invalid date format for hire_date. Use YYYY-MM-DD'}), 400)
                return add_cors_headers(response)
        
        # 新しい社員を作成
        employee = Employee(
            name=data.get('name'),
            department=data.get('department'),
            position=data.get('position'),
            email=data.get('email'),
            phone_number=data.get('phone_number'),
            hire_date=hire_date
        )
        
        db.session.add(employee)
        db.session.commit()
        
        app.logger.info(f"社員追加成功 ID: {employee.id}")
        response = make_response(jsonify({
            'message': 'Employee added successfully',
            'employee': employee.to_dict()
        }), 201)
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/employees/<int:employee_id>', methods=['PUT'])
def update_employee(employee_id):
    try:
        employee = Employee.query.get(employee_id)
        if not employee:
            response = make_response(jsonify({'error': 'Employee not found'}), 404)
            return add_cors_headers(response)
        
        data = request.json
        app.logger.info(f"社員更新リクエスト ID: {employee_id}, データ: {data}")
        
        # 日付文字列をDateオブジェクトに変換
        if 'hire_date' in data and data['hire_date']:
            hire_date = parse_date(data['hire_date'])
            if hire_date is None:
                response = make_response(jsonify({'error': 'Invalid date format for hire_date. Use YYYY-MM-DD'}), 400)
                return add_cors_headers(response)
            employee.hire_date = hire_date
        elif 'hire_date' in data and data['hire_date'] is None:
            employee.hire_date = None
        
        # 各フィールドを更新
        for field in ['name', 'department', 'position', 'email', 'phone_number']:
            if field in data:
                setattr(employee, field, data[field])
        
        db.session.commit()
        app.logger.info(f"社員更新成功 ID: {employee_id}")
        response = make_response(jsonify({
            'message': 'Employee updated successfully',
            'employee': employee.to_dict()
        }))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/employees/<int:employee_id>', methods=['DELETE'])
def delete_employee(employee_id):
    try:
        employee = Employee.query.get(employee_id)
        if not employee:
            response = make_response(jsonify({'error': 'Employee not found'}), 404)
            return add_cors_headers(response)
        
        db.session.delete(employee)
        db.session.commit()
        app.logger.info(f"社員削除成功 ID: {employee_id}")
        response = make_response(jsonify({'message': 'Employee deleted successfully'}))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/applicants/<int:applicant_id>/assign-employee', methods=['PUT'])
def assign_employee_to_applicant(applicant_id):
    try:
        applicant = Applicant.query.get(applicant_id)
        if not applicant:
            response = make_response(jsonify({'error': 'Applicant not found'}), 404)
            return add_cors_headers(response)

        data = request.json
        app.logger.info(f"求職者担当社員更新リクエスト ID: {applicant_id}, データ: {data}")

        # 担当社員IDが提供されているか確認
        if 'employee_id' not in data:
            response = make_response(jsonify({'error': 'Employee ID is required'}), 400)
            return add_cors_headers(response)

        # 空の文字列が送られてきた場合はNoneに設定
        employee_id = data['employee_id'] if data['employee_id'] else None

        # 指定された社員が存在するか確認（空の場合はスキップ）
        if employee_id:
            employee = Employee.query.get(employee_id)
            if not employee:
                response = make_response(jsonify({'error': 'Specified employee not found'}), 400)
                return add_cors_headers(response)

        # 担当社員IDを設定
        applicant.assigned_employee_id = employee_id
        
        # データベースを更新
        db.session.commit()
        app.logger.info(f"求職者担当社員更新成功 ID: {applicant_id}, 担当社員ID: {employee_id}")

        # 更新後の求職者データを取得して返す
        updated_applicant_data = applicant.to_dict() 
        updated_applicant_data['assigned_employee_id'] = employee_id
        updated_applicant_data['employee_id'] = employee_id  # 互換性のため両方設定
        
        response = make_response(jsonify({
            'message': 'Assigned employee updated successfully',
            'applicant_id': applicant_id,
            'employee_id': employee_id,
            'applicant': updated_applicant_data  # 完全な更新済み求職者データを返す
        }))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)


# 社員KPI APIエンドポイント
@app.route('/api/employees/<int:employee_id>/kpi', methods=['GET'])
def get_employee_kpi(employee_id):
    try:
        # 指定された社員が存在するか確認
        employee = Employee.query.get(employee_id)
        if not employee:
            response = make_response(jsonify({'error': 'Employee not found'}), 404)
            return add_cors_headers(response)
        
        # タイムフレームの取得（デフォルトは月間）
        timeframe = request.args.get('timeframe', 'month')
        
        # タイムフレームに基づいて日付範囲を計算
        today = date.today()
        if timeframe == 'week':
            start_date = today - timedelta(days=7)
        elif timeframe == 'month':
            start_date = today.replace(day=1)
        elif timeframe == 'quarter':
            quarter_start_month = ((today.month - 1) // 3) * 3 + 1
            start_date = date(today.year, quarter_start_month, 1)
        elif timeframe == 'year':
            start_date = date(today.year, 1, 1)
        else:
            start_date = today.replace(day=1)  # デフォルトは月間
        
        # 社員に関連する求職者を取得
        # 注意: 実際のデータモデルに応じて、社員と求職者の関連付け方法を調整する必要があります
        # ここではPhoneCallテーブルを介して関連付けられていると仮定
        applicant_ids = db.session.query(PhoneCall.applicant_id).filter(
            PhoneCall.employee_id == employee_id
        ).distinct().all()
        applicant_ids = [id[0] for id in applicant_ids]
        
        # 各ステージの件数を取得
        total_applicants = len(applicant_ids)
        
        # 各ステージの集計を行う
        # 実際のデータモデルに応じて調整が必要
        # phoneCallsテーブルから架電数、接続数を集計
        total_calls = PhoneCall.query.filter(
            PhoneCall.employee_id == employee_id,
            PhoneCall.call_date >= start_date
        ).count()
        
        total_connections = PhoneCall.query.filter(
            PhoneCall.employee_id == employee_id,
            PhoneCall.status == 'completed',
            PhoneCall.call_date >= start_date
        ).count()
        
        # 求職者ステータスから各ステージの数を集計
        # （connection_date, proposal_date, document_sent_date, ...など）
        applicants = Applicant.query.filter(Applicant.id.in_(applicant_ids)).all()
        
        total_proposals = sum(1 for a in applicants if a.proposal_date and a.proposal_date >= start_date)
        total_documents_sent = sum(1 for a in applicants if a.document_sent_date and a.document_sent_date >= start_date)
        total_documents_passed = sum(1 for a in applicants if a.document_passed_date and a.document_passed_date >= start_date)
        total_interviews = sum(1 for a in applicants if a.interview_date and a.interview_date >= start_date)
        total_offers = sum(1 for a in applicants if a.offer_date and a.offer_date >= start_date)
        total_hires = sum(1 for a in applicants if a.hire_date and a.hire_date >= start_date)
        total_payments = sum(1 for a in applicants if a.payment_date and a.payment_date >= start_date)
        
        # 総売上の計算（紹介料の合計）
        total_revenue = sum(a.referral_fee or 0 for a in applicants if a.payment_date and a.payment_date >= start_date)
        
        # 変換率の計算
        call_to_connection = (total_connections / total_calls * 100) if total_calls > 0 else 0
        connection_to_proposal = (total_proposals / total_connections * 100) if total_connections > 0 else 0
        proposal_to_document = (total_documents_sent / total_proposals * 100) if total_proposals > 0 else 0
        document_to_pass = (total_documents_passed / total_documents_sent * 100) if total_documents_sent > 0 else 0
        interview_to_offer = (total_offers / total_interviews * 100) if total_interviews > 0 else 0
        offer_to_hire = (total_hires / total_offers * 100) if total_offers > 0 else 0
        hire_to_payment = (total_payments / total_hires * 100) if total_hires > 0 else 0
        
        # 月次進捗の計算
        # ここでは、過去4ヶ月分のデータを生成
        month_data = []
        for i in range(3, -1, -1):
            if timeframe == 'week':
                # 週間の場合は、過去4週間を取得
                period_start = today - timedelta(days=(i+1)*7)
                period_end = today - timedelta(days=i*7)
                period_name = f"{period_start.strftime('%m/%d')}~{period_end.strftime('%m/%d')}"
            elif timeframe == 'month':
                # 月間の場合は、過去4ヶ月を取得
                month = today.month - i
                year = today.year
                while month <= 0:
                    month += 12
                    year -= 1
                period_start = date(year, month, 1)
                # 次の月の初日から1日引く
                next_month = month + 1
                next_year = year
                if next_month > 12:
                    next_month = 1
                    next_year += 1
                period_end = date(next_year, next_month, 1) - timedelta(days=1)
                period_name = f"{period_start.strftime('%Y/%m')}"
            elif timeframe == 'quarter':
                # 四半期の場合
                quarter = (today.month - 1) // 3 - i
                year = today.year
                while quarter < 0:
                    quarter += 4
                    year -= 1
                period_start = date(year, quarter * 3 + 1, 1)
                # 次の四半期の初日から1日引く
                next_quarter = quarter + 1
                next_year = year
                if next_quarter >= 4:
                    next_quarter = 0
                    next_year += 1
                period_end = date(next_year, next_quarter * 3 + 1, 1) - timedelta(days=1)
                period_name = f"Q{quarter+1}/{year}"
            else:  # 年間の場合
                period_start = date(today.year - i, 1, 1)
                period_end = date(today.year - i, 12, 31)
                period_name = f"{period_start.year}年"
            
            # 各期間ごとのデータを集計
            period_calls = PhoneCall.query.filter(
                PhoneCall.employee_id == employee_id,
                PhoneCall.call_date >= period_start,
                PhoneCall.call_date <= period_end
            ).count()
            
            period_connections = PhoneCall.query.filter(
                PhoneCall.employee_id == employee_id,
                PhoneCall.status == 'completed',
                PhoneCall.call_date >= period_start,
                PhoneCall.call_date <= period_end
            ).count()
            
            period_proposals = sum(1 for a in applicants if a.proposal_date and period_start <= a.proposal_date <= period_end)
            period_documents = sum(1 for a in applicants if a.document_sent_date and period_start <= a.document_sent_date <= period_end)
            period_passes = sum(1 for a in applicants if a.document_passed_date and period_start <= a.document_passed_date <= period_end)
            period_interviews = sum(1 for a in applicants if a.interview_date and period_start <= a.interview_date <= period_end)
            period_offers = sum(1 for a in applicants if a.offer_date and period_start <= a.offer_date <= period_end)
            period_hires = sum(1 for a in applicants if a.hire_date and period_start <= a.hire_date <= period_end)
            period_payments = sum(1 for a in applicants if a.payment_date and period_start <= a.payment_date <= period_end)
            
            # 月間売上の計算
            period_revenue = sum(a.referral_fee or 0 for a in applicants 
                               if a.payment_date and period_start <= a.payment_date <= period_end)
            
            month_data.append({
                'month': period_name,
                'calls': period_calls,
                'connections': period_connections,
                'proposals': period_proposals,
                'documents': period_documents,
                'passes': period_passes,
                'interviews': period_interviews,
                'offers': period_offers,
                'hires': period_hires,
                'payments': period_payments,
                'revenue': period_revenue
            })
        
        # パイプライン分布の計算
        # 現在のステータスに基づいて求職者の分布を集計
        pipeline_distribution = []
        
        # 架電待ち（申込のみで架電がないケース）
        call_waiting = sum(1 for a in applicants 
                          if not PhoneCall.query.filter(PhoneCall.applicant_id == a.id).first())
        pipeline_distribution.append({
            'name': '架電待ち',
            'value': call_waiting
        })
        
        # 接続待ち（架電はあるが接続されていないケース）
        connection_waiting = sum(1 for a in applicants 
                               if PhoneCall.query.filter(
                                   PhoneCall.applicant_id == a.id, 
                                   PhoneCall.status != 'completed'
                               ).first())
        pipeline_distribution.append({
            'name': '接続待ち',
            'value': connection_waiting
        })
        
        # 提案済み（接続後、提案日があるが書類送付日がないケース）
        proposal_done = sum(1 for a in applicants 
                           if a.proposal_date and not a.document_sent_date)
        pipeline_distribution.append({
            'name': '提案済み',
            'value': proposal_done
        })
        
        # 書類送付済み（書類送付日があるが通過日がないケース）
        document_sent = sum(1 for a in applicants 
                           if a.document_sent_date and not a.document_passed_date)
        pipeline_distribution.append({
            'name': '書類送付済み',
            'value': document_sent
        })
        
        # 選考通過（書類通過日があるが面接日がないケース）
        document_passed = sum(1 for a in applicants 
                             if a.document_passed_date and not a.interview_date)
        pipeline_distribution.append({
            'name': '選考通過',
            'value': document_passed
        })
        
        # 面接調整中（面接日があるが内定日がないケース）
        interview_scheduled = sum(1 for a in applicants 
                                 if a.interview_date and not a.offer_date)
        pipeline_distribution.append({
            'name': '面接調整中',
            'value': interview_scheduled
        })
        
        # 内定待ち（内定日があるが入社日がないケース）
        offer_waiting = sum(1 for a in applicants 
                           if a.offer_date and not a.hire_date)
        pipeline_distribution.append({
            'name': '内定待ち',
            'value': offer_waiting
        })
        
        # 入社待ち（入社日があるが入金日がないケース）
        hire_waiting = sum(1 for a in applicants 
                          if a.hire_date and not a.payment_date)
        pipeline_distribution.append({
            'name': '入社待ち',
            'value': hire_waiting
        })
        
        # 入金待ち（入社日があり今後入金日が設定される予定のケース）
        payment_waiting = sum(1 for a in applicants 
                             if a.hire_date and not a.payment_date)
        pipeline_distribution.append({
            'name': '入金待ち',
            'value': payment_waiting
        })
        
        # ステージ間の平均日数の計算
        time_between_stages = []
        
        # 架電→接続
        connection_days = []
        for a in applicants:
            # 最初の架電日を取得
            first_call = PhoneCall.query.filter(
                PhoneCall.applicant_id == a.id
            ).order_by(PhoneCall.call_date).first()
            
            # 最初の接続日を取得
            first_connection = PhoneCall.query.filter(
                PhoneCall.applicant_id == a.id,
                PhoneCall.status == 'completed'
            ).order_by(PhoneCall.call_date).first()
            
            if first_call and first_connection and first_call.call_date < first_connection.call_date:
                days = (first_connection.call_date - first_call.call_date).days
                connection_days.append(days)
        
        avg_connection_days = sum(connection_days) / len(connection_days) if connection_days else 0
        time_between_stages.append({
            'name': '架電→接続',
            'days': avg_connection_days
        })
        
        # 接続→提案
        proposal_days = []
        for a in applicants:
            # 最初の接続日を取得
            first_connection = PhoneCall.query.filter(
                PhoneCall.applicant_id == a.id,
                PhoneCall.status == 'completed'
            ).order_by(PhoneCall.call_date).first()
            
            if first_connection and a.proposal_date and first_connection.call_date.date() < a.proposal_date:
                days = (a.proposal_date - first_connection.call_date.date()).days
                proposal_days.append(days)
        
        avg_proposal_days = sum(proposal_days) / len(proposal_days) if proposal_days else 0
        time_between_stages.append({
            'name': '接続→提案',
            'days': avg_proposal_days
        })
        
        # 提案→書類送付
        document_days = []
        for a in applicants:
            if a.proposal_date and a.document_sent_date and a.proposal_date < a.document_sent_date:
                days = (a.document_sent_date - a.proposal_date).days
                document_days.append(days)
        
        avg_document_days = sum(document_days) / len(document_days) if document_days else 0
        time_between_stages.append({
            'name': '提案→書類送付',
            'days': avg_document_days
        })
        
        # 書類送付→通過
        pass_days = []
        for a in applicants:
            if a.document_sent_date and a.document_passed_date and a.document_sent_date < a.document_passed_date:
                days = (a.document_passed_date - a.document_sent_date).days
                pass_days.append(days)
        
        avg_pass_days = sum(pass_days) / len(pass_days) if pass_days else 0
        time_between_stages.append({
            'name': '書類送付→通過',
            'days': avg_pass_days
        })
        
        # 通過→面接
        interview_days = []
        for a in applicants:
            if a.document_passed_date and a.interview_date and a.document_passed_date < a.interview_date:
                days = (a.interview_date - a.document_passed_date).days
                interview_days.append(days)
        
        avg_interview_days = sum(interview_days) / len(interview_days) if interview_days else 0
        time_between_stages.append({
            'name': '通過→面接',
            'days': avg_interview_days
        })
        
        # 面接→内定
        offer_days = []
        for a in applicants:
            if a.interview_date and a.offer_date and a.interview_date < a.offer_date:
                days = (a.offer_date - a.interview_date).days
                offer_days.append(days)
        
        avg_offer_days = sum(offer_days) / len(offer_days) if offer_days else 0
        time_between_stages.append({
            'name': '面接→内定',
            'days': avg_offer_days
        })
        
        # 内定→入社
        hire_days = []
        for a in applicants:
            if a.offer_date and a.hire_date and a.offer_date < a.hire_date:
                days = (a.hire_date - a.offer_date).days
                hire_days.append(days)
        
        avg_hire_days = sum(hire_days) / len(hire_days) if hire_days else 0
        time_between_stages.append({
            'name': '内定→入社',
            'days': avg_hire_days
        })
        
        # 入社→入金
        payment_days = []
        for a in applicants:
            if a.hire_date and a.payment_date and a.hire_date < a.payment_date:
                days = (a.payment_date - a.hire_date).days
                payment_days.append(days)
        
        avg_payment_days = sum(payment_days) / len(payment_days) if payment_days else 0
        time_between_stages.append({
            'name': '入社→入金',
            'days': avg_payment_days
        })
        
        # 結果をJSONで返す
        result = {
            'name': employee.name,
            'department': employee.department,
            'position': employee.position,
            
            # サマリー指標
            'summary': {
                'totalApplicants': total_applicants,
                'totalCalls': total_calls,
                'totalConnections': total_connections,
                'totalProposals': total_proposals,
                'totalDocumentsSent': total_documents_sent,
                'totalDocumentsPassed': total_documents_passed,
                'totalInterviews': total_interviews,
                'totalOffers': total_offers,
                'totalHires': total_hires,
                'totalPayments': total_payments,
                'totalRevenue': total_revenue,
            },
            
            # 変換率
            'conversionRates': {
                'callToConnection': call_to_connection,
                'connectionToProposal': connection_to_proposal,
                'proposalToDocument': proposal_to_document,
                'documentToPass': document_to_pass,
                'interviewToOffer': interview_to_offer,
                'offerToHire': offer_to_hire,
                'hireToPayment': hire_to_payment,
            },
            
            # 月次進捗
            'monthlyProgress': month_data,
            
            # パイプライン分布
            'pipelineDistribution': pipeline_distribution,
            
            # ステージ間の平均日数
            'timeBetweenStages': time_between_stages,
        }
        
        response = make_response(jsonify(result))
        return add_cors_headers(response)
        
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# 会社全体KPI APIエンドポイント
@app.route('/api/company/kpi', methods=['GET'])
def get_company_kpi():
    try:
        # タイムフレームの取得（デフォルトは月間）
        timeframe = request.args.get('timeframe', 'month')
        
        # タイムフレームに基づいて日付範囲を計算
        today = date.today()
        if timeframe == 'week':
            start_date = today - timedelta(days=7)
        elif timeframe == 'month':
            start_date = today.replace(day=1)
        elif timeframe == 'quarter':
            quarter_start_month = ((today.month - 1) // 3) * 3 + 1
            start_date = date(today.year, quarter_start_month, 1)
        elif timeframe == 'year':
            start_date = date(today.year, 1, 1)
        else:
            start_date = today.replace(day=1)  # デフォルトは月間
        
        # 全求職者を取得
        applicants = Applicant.query.all()
        total_applicants = len(applicants)
        
        # 全社員を取得
        employees = Employee.query.all()
        
        # 架電数の集計
        total_calls = PhoneCall.query.filter(
            PhoneCall.call_date >= start_date
        ).count()
        
        # 接続数の集計
        total_connections = PhoneCall.query.filter(
            PhoneCall.status == 'completed',
            PhoneCall.call_date >= start_date
        ).count()
        
        # 各ステージの数を集計
        total_proposals = sum(1 for a in applicants if a.proposal_date and a.proposal_date >= start_date)
        total_documents_sent = sum(1 for a in applicants if a.document_sent_date and a.document_sent_date >= start_date)
        total_documents_passed = sum(1 for a in applicants if a.document_passed_date and a.document_passed_date >= start_date)
        total_interviews = sum(1 for a in applicants if a.interview_date and a.interview_date >= start_date)
        total_offers = sum(1 for a in applicants if a.offer_date and a.offer_date >= start_date)
        total_hires = sum(1 for a in applicants if a.hire_date and a.hire_date >= start_date)
        total_payments = sum(1 for a in applicants if a.payment_date and a.payment_date >= start_date)
        
        # 総売上の計算
        total_revenue = sum(a.referral_fee or 0 for a in applicants if a.payment_date and a.payment_date >= start_date)
        
        # 変換率の計算
        call_to_connection = (total_connections / total_calls * 100) if total_calls > 0 else 0
        connection_to_proposal = (total_proposals / total_connections * 100) if total_connections > 0 else 0
        proposal_to_document = (total_documents_sent / total_proposals * 100) if total_proposals > 0 else 0
        document_to_pass = (total_documents_passed / total_documents_sent * 100) if total_documents_sent > 0 else 0
        interview_to_offer = (total_offers / total_interviews * 100) if total_interviews > 0 else 0
        offer_to_hire = (total_hires / total_offers * 100) if total_offers > 0 else 0
        hire_to_payment = (total_payments / total_hires * 100) if total_hires > 0 else 0
        
        # 全体変換率の計算（応募から入社までの比率）
        overall_conversion_rate = (total_hires / total_applicants * 100) if total_applicants > 0 else 0
        
        # 平均採用期間の計算
        hiring_periods = []
        for a in applicants:
            if a.application_date and a.hire_date and a.application_date < a.hire_date:
                days = (a.hire_date - a.application_date).days
                hiring_periods.append(days)
        
        average_time_to_hire = sum(hiring_periods) / len(hiring_periods) if hiring_periods else 0
        
        # 月次進捗の計算
        month_data = []
        for i in range(3, -1, -1):
            if timeframe == 'week':
                # 週間の場合は、過去4週間を取得
                period_start = today - timedelta(days=(i+1)*7)
                period_end = today - timedelta(days=i*7)
                period_name = f"{period_start.strftime('%m/%d')}~{period_end.strftime('%m/%d')}"
            elif timeframe == 'month':
                # 月間の場合は、過去4ヶ月を取得
                month = today.month - i
                year = today.year
                while month <= 0:
                    month += 12
                    year -= 1
                period_start = date(year, month, 1)
                # 次の月の初日から1日引く
                next_month = month + 1
                next_year = year
                if next_month > 12:
                    next_month = 1
                    next_year += 1
                period_end = date(next_year, next_month, 1) - timedelta(days=1)
                period_name = f"{period_start.strftime('%Y/%m')}"
            elif timeframe == 'quarter':
                # 四半期の場合
                quarter = (today.month - 1) // 3 - i
                year = today.year
                while quarter < 0:
                    quarter += 4
                    year -= 1
                period_start = date(year, quarter * 3 + 1, 1)
                # 次の四半期の初日から1日引く
                next_quarter = quarter + 1
                next_year = year
                if next_quarter >= 4:
                    next_quarter = 0
                    next_year += 1
                period_end = date(next_year, next_quarter * 3 + 1, 1) - timedelta(days=1)
                period_name = f"Q{quarter+1}/{year}"
            else:  # 年間の場合
                period_start = date(today.year - i, 1, 1)
                period_end = date(today.year - i, 12, 31)
                period_name = f"{period_start.year}年"
            
            # 各期間ごとのデータを集計
            period_calls = PhoneCall.query.filter(
                PhoneCall.call_date >= period_start,
                PhoneCall.call_date <= period_end
            ).count()
            
            period_connections = PhoneCall.query.filter(
                PhoneCall.status == 'completed',
                PhoneCall.call_date >= period_start,
                PhoneCall.call_date <= period_end
            ).count()
            
            period_proposals = sum(1 for a in applicants if a.proposal_date and period_start <= a.proposal_date <= period_end)
            period_documents = sum(1 for a in applicants if a.document_sent_date and period_start <= a.document_sent_date <= period_end)
            period_passes = sum(1 for a in applicants if a.document_passed_date and period_start <= a.document_passed_date <= period_end)
            period_interviews = sum(1 for a in applicants if a.interview_date and period_start <= a.interview_date <= period_end)
            period_offers = sum(1 for a in applicants if a.offer_date and period_start <= a.offer_date <= period_end)
            period_hires = sum(1 for a in applicants if a.hire_date and period_start <= a.hire_date <= period_end)
            period_payments = sum(1 for a in applicants if a.payment_date and period_start <= a.payment_date <= period_end)
            
            # 月間売上の計算
            period_revenue = sum(a.referral_fee or 0 for a in applicants 
                               if a.payment_date and period_start <= a.payment_date <= period_end)
            
            month_data.append({
                'month': period_name,
                'calls': period_calls,
                'connections': period_connections,
                'proposals': period_proposals,
                'documents': period_documents,
                'passes': period_passes,
                'interviews': period_interviews,
                'offers': period_offers,
                'hires': period_hires,
                'payments': period_payments,
                'revenue': period_revenue
            })
        
        # パイプライン分布の計算（会社全体）
        pipeline_distribution = []
        
        # 架電待ち（申込のみで架電がないケース）
        call_waiting = sum(1 for a in applicants 
                          if not PhoneCall.query.filter(PhoneCall.applicant_id == a.id).first())
        pipeline_distribution.append({
            'name': '架電待ち',
            'value': call_waiting
        })
        
        # 接続待ち（架電はあるが接続されていないケース）
        connection_waiting = sum(1 for a in applicants 
                               if PhoneCall.query.filter(
                                   PhoneCall.applicant_id == a.id, 
                                   PhoneCall.status != 'completed'
                               ).first())
        pipeline_distribution.append({
            'name': '接続待ち',
            'value': connection_waiting
        })
        
        # 提案済み（接続後、提案日があるが書類送付日がないケース）
        proposal_done = sum(1 for a in applicants 
                           if a.proposal_date and not a.document_sent_date)
        pipeline_distribution.append({
            'name': '提案済み',
            'value': proposal_done
        })
        
        # 書類送付済み（書類送付日があるが通過日がないケース）
        document_sent = sum(1 for a in applicants 
                           if a.document_sent_date and not a.document_passed_date)
        pipeline_distribution.append({
            'name': '書類送付済み',
            'value': document_sent
        })
        
        # 選考通過（書類通過日があるが面接日がないケース）
        document_passed = sum(1 for a in applicants 
                             if a.document_passed_date and not a.interview_date)
        pipeline_distribution.append({
            'name': '選考通過',
            'value': document_passed
        })
        
        # 面接調整中（面接日があるが内定日がないケース）
        interview_scheduled = sum(1 for a in applicants 
                                 if a.interview_date and not a.offer_date)
        pipeline_distribution.append({
            'name': '面接調整中',
            'value': interview_scheduled
        })
        
        # 内定待ち（内定日があるが入社日がないケース）
        offer_waiting = sum(1 for a in applicants 
                           if a.offer_date and not a.hire_date)
        pipeline_distribution.append({
            'name': '内定待ち',
            'value': offer_waiting
        })
        
        # 入社待ち（入社日があるが入金日がないケース）
        hire_waiting = sum(1 for a in applicants 
                          if a.hire_date and not a.payment_date)
        pipeline_distribution.append({
            'name': '入社待ち',
            'value': hire_waiting
        })
        
        # 入金待ち（入社日があり今後入金日が設定される予定のケース）
        payment_waiting = sum(1 for a in applicants 
                             if a.hire_date and not a.payment_date)
        pipeline_distribution.append({
            'name': '入金待ち',
            'value': payment_waiting
        })
        
        # 部門別成績の集計
        department_performance = []
        
        # 部門ごとにグループ化
        departments = {}
        for employee in employees:
            if employee.department not in departments:
                departments[employee.department] = []
            departments[employee.department].append(employee.id)
        
        # 部門ごとの実績を集計
        for department, emp_ids in departments.items():
            # 部門の社員に関連する求職者を取得
            department_applicant_ids = db.session.query(PhoneCall.applicant_id).filter(
                PhoneCall.employee_id.in_(emp_ids)
            ).distinct().all()
            department_applicant_ids = [id[0] for id in department_applicant_ids]
            
            # 入社数を集計
            department_hires = sum(1 for a in Applicant.query.filter(
                Applicant.id.in_(department_applicant_ids),
                Applicant.hire_date != None,
                Applicant.hire_date >= start_date
            ).all())
            
            # 売上を集計
            department_revenue = sum(a.referral_fee or 0 for a in Applicant.query.filter(
                Applicant.id.in_(department_applicant_ids),
                Applicant.payment_date != None,
                Applicant.payment_date >= start_date
            ).all())
            
            department_performance.append({
                'department': department or '未分類',
                'hires': department_hires,
                'revenue': department_revenue
            })
        
        # 四半期ごとの目標vs実績
        # ここではデモのための仮データを生成
        # 実際のアプリケーションでは目標値をデータベースから取得する必要がある
        quarterly_performance = [
            {'quarter': 'Q1', 'target': 42, 'actual': 45},
            {'quarter': 'Q2', 'target': 48, 'actual': 52},
            {'quarter': 'Q3', 'target': 45, 'actual': 38},
            {'quarter': 'Q4', 'target': 40, 'actual': 15}  # 進行中
        ]
        
        # 結果をJSONで返す
        result = {
            # サマリー指標
            'summary': {
                'totalApplicants': total_applicants,
                'totalCalls': total_calls,
                'totalConnections': total_connections,
                'totalProposals': total_proposals,
                'totalDocumentsSent': total_documents_sent,
                'totalDocumentsPassed': total_documents_passed,
                'totalInterviews': total_interviews,
                'totalOffers': total_offers,
                'totalHires': total_hires,
                'totalPayments': total_payments,
                'totalRevenue': total_revenue,
                'conversionRate': overall_conversion_rate,
                'averageTimeToHire': average_time_to_hire,
            },
            
            # 変換率
            'conversionRates': {
                'callToConnection': call_to_connection,
                'connectionToProposal': connection_to_proposal,
                'proposalToDocument': proposal_to_document,
                'documentToPass': document_to_pass,
                'interviewToOffer': interview_to_offer,
                'offerToHire': offer_to_hire,
                'hireToPayment': hire_to_payment,
            },
            
            # 月次進捗
            'monthlyProgress': month_data,
            
            # パイプライン分布
            'pipelineDistribution': pipeline_distribution,
            
            # 部門別成績
            'departmentPerformance': department_performance,
            
            # 四半期目標vs実績
            'quarterlyPerformance': quarterly_performance,
        }
        
        response = make_response(jsonify(result))
        return add_cors_headers(response)
        
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# トップパフォーマー取得APIエンドポイント
@app.route('/api/company/top-performers', methods=['GET'])
def get_top_performers():
    try:
        # タイムフレームの取得（デフォルトは月間）
        timeframe = request.args.get('timeframe', 'month')
        
        # タイムフレームに基づいて日付範囲を計算
        today = date.today()
        if timeframe == 'week':
            start_date = today - timedelta(days=7)
        elif timeframe == 'month':
            start_date = today.replace(day=1)
        elif timeframe == 'quarter':
            quarter_start_month = ((today.month - 1) // 3) * 3 + 1
            start_date = date(today.year, quarter_start_month, 1)
        elif timeframe == 'year':
            start_date = date(today.year, 1, 1)
        else:
            start_date = today.replace(day=1)  # デフォルトは月間
        
        # 各社員のパフォーマンスを計算
        employee_performances = []
        
        for employee in Employee.query.all():
            # 社員に関連する求職者を取得
            applicant_ids = db.session.query(PhoneCall.applicant_id).filter(
                PhoneCall.employee_id == employee.id
            ).distinct().all()
            applicant_ids = [id[0] for id in applicant_ids]
            
            # 入社数を計算
            hires = sum(1 for a in Applicant.query.filter(
                Applicant.id.in_(applicant_ids),
                Applicant.hire_date != None,
                Applicant.hire_date >= start_date
            ).all())
            
            # 売上を計算
            revenue = sum(a.referral_fee or 0 for a in Applicant.query.filter(
                Applicant.id.in_(applicant_ids),
                Applicant.payment_date != None,
                Applicant.payment_date >= start_date
            ).all())
            
            employee_performances.append({
                'id': employee.id,
                'name': employee.name,
                'department': employee.department,
                'hires': hires,
                'revenue': revenue
            })
        
        # 売上ベースでソート（降順）
        employee_performances.sort(key=lambda x: x['revenue'], reverse=True)
        
        # 上位5名を取得
        top_performers = employee_performances[:5]
        
        response = make_response(jsonify(top_performers))
        return add_cors_headers(response)
        
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)


# 面接関連のAPI
@app.route('/api/interviews', methods=['GET'])
def get_interviews():
    try:
        interviews = Interview.query.all()
        response = make_response(jsonify([interview.to_dict() for interview in interviews]))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/interviews', methods=['POST'])
def add_interview():
    try:
        data = request.json
        app.logger.info(f"面接追加リクエスト: {data}")
        
        # 日付文字列をDatetimeオブジェクトに変換
        interview_date = None
        if 'date' in data and data['date']:
            try:
                interview_date = datetime.fromisoformat(data['date'].replace('Z', '+00:00'))
            except ValueError:
                response = make_response(jsonify({'error': 'Invalid date format for interview date. Use ISO format'}), 400)
                return add_cors_headers(response)
        
        # 新しい面接を作成
        interview = Interview(
            applicant_id=data.get('applicant_id'),
            job_id=data.get('job_id'),
            date=interview_date,
            status=data.get('status', 'scheduled'),
            result=data.get('result'),
            notes=data.get('notes'),
            preparation_info=data.get('preparation_info')
        )
        
        db.session.add(interview)
        db.session.commit()
        app.logger.info(f"面接追加成功 ID: {interview.id}")
        response = make_response(jsonify({
            'message': 'Interview added successfully',
            'interview': interview.to_dict()
        }), 201)
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/interviews/<int:interview_id>', methods=['GET'])
def get_interview(interview_id):
    try:
        interview = Interview.query.get(interview_id)
        if interview:
            response = make_response(jsonify(interview.to_dict()))
            return add_cors_headers(response)
        response = make_response(jsonify({'error': 'Interview not found'}), 404)
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/interviews/<int:interview_id>', methods=['PUT'])
def update_interview(interview_id):
    try:
        interview = Interview.query.get(interview_id)
        if not interview:
            response = make_response(jsonify({'error': 'Interview not found'}), 404)
            return add_cors_headers(response)
        
        data = request.json
        app.logger.info(f"面接更新リクエスト ID: {interview_id}, データ: {data}")
        
        # 日時文字列をDatetimeオブジェクトに変換
        if 'date' in data and data['date']:
            try:
                interview.date = datetime.fromisoformat(data['date'].replace('Z', '+00:00'))
            except ValueError:
                response = make_response(jsonify({'error': 'Invalid date format for interview date. Use ISO format'}), 400)
                return add_cors_headers(response)
        
        # 各フィールドを更新
        for field in ['applicant_id', 'job_id', 'status', 'result', 'notes', 'preparation_info']:
            if field in data:
                setattr(interview, field, data[field])
        
        db.session.commit()
        app.logger.info(f"面接更新成功 ID: {interview_id}")
        response = make_response(jsonify({
            'message': 'Interview updated successfully',
            'interview': interview.to_dict()
        }))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/interviews/<int:interview_id>', methods=['DELETE'])
def delete_interview(interview_id):
    try:
        interview = Interview.query.get(interview_id)
        if not interview:
            response = make_response(jsonify({'error': 'Interview not found'}), 404)
            return add_cors_headers(response)
        
        db.session.delete(interview)
        db.session.commit()
        app.logger.info(f"面接削除成功 ID: {interview_id}")
        response = make_response(jsonify({'message': 'Interview deleted successfully'}))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# 統計分析用のAPI
@app.route('/api/statistics/interview-results', methods=['GET'])
def get_interview_statistics():
    try:
        # 面接結果の統計
        passed = Interview.query.filter_by(result='passed').count()
        failed = Interview.query.filter_by(result='failed').count()
        total = Interview.query.count()
        pass_rate = (passed / total) * 100 if total > 0 else 0
        
        response = make_response(jsonify({
            'total_interviews': total,
            'passed': passed,
            'failed': failed,
            'pass_rate': pass_rate
        }))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# 架電記録関連のAPI
@app.route('/api/phone-calls', methods=['GET'])
def get_phone_calls():
    try:
        phone_calls = PhoneCall.query.all()
        response = make_response(jsonify([phone_call.to_dict() for phone_call in phone_calls]))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/phone-calls/<int:phone_call_id>', methods=['GET'])
def get_phone_call(phone_call_id):
    try:
        phone_call = PhoneCall.query.get(phone_call_id)
        if phone_call:
            response = make_response(jsonify(phone_call.to_dict()))
            return add_cors_headers(response)
        response = make_response(jsonify({'error': 'Phone call record not found'}), 404)
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/phone-calls', methods=['POST'])
def add_phone_call():
    try:
        data = request.json
        app.logger.info(f"架電記録追加リクエスト: {data}")
        
        # 必須フィールドの検証
        if not data.get('applicant_id'):
            response = make_response(jsonify({'error': 'Applicant ID is required'}), 400)
            return add_cors_headers(response)
        
        # 日時文字列をDatetimeオブジェクトに変換
        call_date = datetime.utcnow()
        if 'call_date' in data and data['call_date']:
            try:
                call_date = datetime.fromisoformat(data['call_date'].replace('Z', '+00:00'))
            except ValueError:
                response = make_response(jsonify({'error': 'Invalid date format for call date. Use ISO format'}), 400)
                return add_cors_headers(response)
        
        follow_up_date = None
        if 'follow_up_date' in data and data['follow_up_date']:
            try:
                follow_up_date = datetime.fromisoformat(data['follow_up_date'].replace('Z', '+00:00'))
            except ValueError:
                response = make_response(jsonify({'error': 'Invalid date format for follow-up date. Use ISO format'}), 400)
                return add_cors_headers(response)
        
        # 新しい架電記録を作成
        phone_call = PhoneCall(
            applicant_id=data.get('applicant_id'),
            employee_id=data.get('employee_id'),
            call_date=call_date,
            status=data.get('status', 'scheduled'),
            notes=data.get('notes'),
            follow_up_date=follow_up_date
        )
        
        db.session.add(phone_call)
        db.session.commit()
        app.logger.info(f"架電記録追加成功 ID: {phone_call.id}")
        response = make_response(jsonify({
            'message': 'Phone call record added successfully',
            'phone_call': phone_call.to_dict()
        }), 201)
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/phone-calls/<int:phone_call_id>', methods=['PUT'])
def update_phone_call(phone_call_id):
    try:
        phone_call = PhoneCall.query.get(phone_call_id)
        if not phone_call:
            response = make_response(jsonify({'error': 'Phone call record not found'}), 404)
            return add_cors_headers(response)
        
        data = request.json
        app.logger.info(f"架電記録更新リクエスト ID: {phone_call_id}, データ: {data}")
        
        # 日時文字列をDatetimeオブジェクトに変換
        if 'call_date' in data and data['call_date']:
            try:
                phone_call.call_date = datetime.fromisoformat(data['call_date'].replace('Z', '+00:00'))
            except ValueError:
                response = make_response(jsonify({'error': 'Invalid date format for call date. Use ISO format'}), 400)
                return add_cors_headers(response)
        
        if 'follow_up_date' in data:
            if data['follow_up_date']:
                try:
                    phone_call.follow_up_date = datetime.fromisoformat(data['follow_up_date'].replace('Z', '+00:00'))
                except ValueError:
                    response = make_response(jsonify({'error': 'Invalid date format for follow-up date. Use ISO format'}), 400)
                    return add_cors_headers(response)
            else:
                phone_call.follow_up_date = None
        
        # 各フィールドを更新
        for field in ['applicant_id', 'employee_id', 'status', 'notes']:
            if field in data:
                setattr(phone_call, field, data[field])
        
        db.session.commit()
        app.logger.info(f"架電記録更新成功 ID: {phone_call_id}")
        response = make_response(jsonify({
            'message': 'Phone call record updated successfully',
            'phone_call': phone_call.to_dict()
        }))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/phone-calls/<int:phone_call_id>', methods=['DELETE'])
def delete_phone_call(phone_call_id):
    try:
        phone_call = PhoneCall.query.get(phone_call_id)
        if not phone_call:
            response = make_response(jsonify({'error': 'Phone call record not found'}), 404)
            return add_cors_headers(response)
        
        db.session.delete(phone_call)
        db.session.commit()
        app.logger.info(f"架電記録削除成功 ID: {phone_call_id}")
        response = make_response(jsonify({'message': 'Phone call record deleted successfully'}))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        db.session.rollback()
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/applicants/<int:applicant_id>/phone-calls', methods=['GET'])
def get_applicant_phone_calls(applicant_id):
    try:
        applicant = Applicant.query.get(applicant_id)
        if not applicant:
            response = make_response(jsonify({'error': 'Applicant not found'}), 404)
            return add_cors_headers(response)
        
        phone_calls = PhoneCall.query.filter_by(applicant_id=applicant_id).all()
        response = make_response(jsonify([phone_call.to_dict() for phone_call in phone_calls]))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# 設定関連のAPI
@app.route('/api/settings', methods=['GET'])
def get_settings():
    try:
        # 設定ファイルが存在するか確認
        settings_file = os.path.join(os.path.dirname(__file__), 'settings.json')
        if not os.path.exists(settings_file):
            # デフォルト設定を作成
            default_settings = {
                'spreadsheet': {
                    'spreadsheet_id': '',
                    'auto_import': False,
                    'import_frequency': 'daily'
                },
                'excel': {
                    'auto_import': False,
                    'import_frequency': 'daily',
                    'last_import_directory': ''
                },
                'notifications': {
                    'email_notifications': False,
                    'app_notifications': True,
                    'email_address': ''
                },
                'ui': {
                    'theme': 'light',
                    'items_per_page': 10,
                    'default_sort': 'created_at'
                },
                'backup': {
                    'auto_backup': False,
                    'backup_frequency': 'weekly',
                    'backup_directory': ''
                }
            }
            
            # デフォルト設定をファイルに保存
            with open(settings_file, 'w') as f:
                json.dump(default_settings, f, indent=2)
            
            response = make_response(jsonify(default_settings))
            return add_cors_headers(response)
        
        # 既存の設定ファイルを読み込む
        with open(settings_file, 'r') as f:
            settings = json.load(f)
        
        response = make_response(jsonify(settings))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

@app.route('/api/settings', methods=['PUT'])
def update_settings():
    try:
        data = request.json
        app.logger.info("設定更新リクエスト受信")
        
        # 設定ファイルのパス
        settings_file = os.path.join(os.path.dirname(__file__), 'settings.json')
        
        # 設定を保存
        with open(settings_file, 'w') as f:
            json.dump(data, f, indent=2)
        
        app.logger.info("設定更新成功")
        response = make_response(jsonify({
            'message': 'Settings updated successfully',
            'settings': data
        }))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# データベースのリセット用エンドポイント
@app.route('/api/reset-database', methods=['POST'])
def reset_database():
    """データベースをリセットするエンドポイント（開発環境のみ）"""
    if not DEBUG:
        response = make_response(jsonify({'error': 'This endpoint is only available in development mode'}), 403)
        return add_cors_headers(response)
    
    try:
        db.drop_all()
        db.create_all()
        app.logger.warning("データベースを初期化しました")
        response = make_response(jsonify({'message': 'Database has been reset successfully'}))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(f"データベースリセットエラー: {str(e)}")
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# 健全性チェックエンドポイント
@app.route('/api/health', methods=['GET'])
def health_check():
    try:
        # データベース接続をチェック
        db.session.execute('SELECT 1').scalar()
        response = make_response(jsonify({'status': 'healthy', 'message': 'API server is running and database connection is OK'}))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(e)
        response = make_response(jsonify({'status': 'unhealthy', 'message': str(e)}), 500)
        return add_cors_headers(response)

# 年齢計算のためのヘルパー関数
def calculate_age(birthdate):
    today = date.today()
    return today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))

# キーワード抽出のヘルパー関数
def extract_keywords(text):
    if not text:
        return []
    # 簡易的なキーワード抽出（実際の実装ではより洗練されたアルゴリズムを使用）
    # 日本語テキストから名詞を抽出するためにMeCabなどを使うとよい
    words = text.lower().replace('、', ' ').replace('，', ' ').replace('。', ' ').split()
    return [w.strip() for w in words if w.strip() and len(w.strip()) > 1]

# マッチングAPIエンドポイント - 修正版 (年齢と希望勤務地のみ考慮)
@app.route('/api/matching/applicant/<int:applicant_id>', methods=['GET', 'OPTIONS'])
def get_applicant_matching_jobs(applicant_id):
    # OPTIONSリクエストの処理
    if request.method == 'OPTIONS':
        app.logger.info(f'OPTIONS リクエスト処理: /api/matching/applicant/{applicant_id}')
        response = make_response()
        return add_cors_headers(response)
    
    try:
        app.logger.info(f'求職者マッチングリクエスト: ID={applicant_id}')
        
        # ソートパラメータを取得 (デフォルトは年齢上限順)
        sort_by = request.args.get('sort_by', 'age_limit')
        app.logger.info(f'ソート条件: {sort_by}')
        
        # 求職者情報を取得
        applicant = Applicant.query.get(applicant_id)
        if not applicant:
            response = make_response(jsonify({'error': 'Applicant not found'}), 404)
            return add_cors_headers(response)
        
        # 求職者の年齢を計算 (誕生日から)
        applicant_age = None
        if applicant.birthdate:
            applicant_age = calculate_age(applicant.birthdate)
            app.logger.info(f'求職者の年齢: {applicant_age}歳')
        
        # 全求人情報を取得
        jobs = Job.query.all()
        app.logger.info(f'求人情報件数: {len(jobs)}件')
        
        # マッチング処理 (年齢と希望勤務地のみ考慮)
        matches = []
        
        for job in jobs:
            match_data = {
                'job': job.to_dict(),
                'age_match': True,
                'location_match': 0
            }
            
            # 1. 年齢条件の確認
            if applicant_age is not None and job.age_limit is not None:
                if applicant_age > job.age_limit:
                    match_data['age_match'] = False
                    app.logger.debug(f'年齢不一致: 求職者={applicant_age}歳, 求人上限={job.age_limit}歳')
            
            # 2. 希望勤務地の一致確認
            if hasattr(applicant, 'desired_location') and applicant.desired_location and job.prefecture:
                # 希望勤務地と都道府県の一致度を評価
                if applicant.desired_location in job.prefecture or job.prefecture in applicant.desired_location:
                    match_data['location_match'] = 100  # 完全一致
                    app.logger.debug(f'勤務地完全一致: {applicant.desired_location} ⇔ {job.prefecture}')
                # 部分一致の確認
                elif any(loc in job.prefecture for loc in extract_keywords(applicant.desired_location)) or \
                     any(loc in applicant.desired_location for loc in extract_keywords(job.prefecture)):
                    match_data['location_match'] = 50  # 部分一致
                    app.logger.debug(f'勤務地部分一致: {applicant.desired_location} ⇔ {job.prefecture}')
            
            # マッチリストに追加
            matches.append(match_data)
        
        # ソート処理
        if sort_by == 'location':
            # 希望勤務地マッチ度でソート（高いものが先）
            app.logger.info('希望勤務地マッチ度順にソート')
            matches = sorted(matches, key=lambda x: (x['location_match']), reverse=True)
        else:
            # デフォルトは年齢条件でソート
            # まず年齢条件を満たすもの優先、次に年齢上限が低い順
            app.logger.info('年齢条件順にソート')
            matches = sorted(matches, key=lambda x: (
                1 if x['age_match'] else 0,  # 年齢条件を満たすものが先
                0 if x['job'].get('age_limit') is None else 1,  # age_limitがないものは後
                -1 * (x['job'].get('age_limit') if x['job'].get('age_limit') is not None else 999)  # 年齢上限が低いものが先
            ), reverse=True)
        
        # 結果を整形
        formatted_matches = []
        for match in matches:
            job_data = match['job']
            formatted_match = {
                'job': job_data,
                'age_match': match['age_match'],
                'location_match': match['location_match'],
                'applicant_age': applicant_age
            }
            formatted_matches.append(formatted_match)
        
        app.logger.info(f'求職者マッチング結果: ID={applicant_id}, 件数={len(formatted_matches)}')
        response = make_response(jsonify(formatted_matches))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(f"求職者マッチングエラー: ID={applicant_id}, エラー={str(e)}")
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

# 逆方向のマッチング - 求人に合う求職者を探す - 修正版 (年齢と希望勤務地のみ考慮)
@app.route('/api/matching/job/<int:job_id>', methods=['GET', 'OPTIONS'])
def get_job_matching_applicants(job_id):
    # OPTIONSリクエストの処理
    if request.method == 'OPTIONS':
        app.logger.info(f'OPTIONS リクエスト処理: /api/matching/job/{job_id}')
        response = make_response()
        return add_cors_headers(response)
    
    try:
        app.logger.info(f'求人マッチングリクエスト: ID={job_id}')
        
        # ソートパラメータを取得
        sort_by = request.args.get('sort_by', 'age')  # デフォルトは年齢順
        
        # 求人情報を取得
        job = Job.query.get(job_id)
        if not job:
            response = make_response(jsonify({'error': 'Job not found'}), 404)
            return add_cors_headers(response)
        
        # 全求職者情報を取得
        applicants = Applicant.query.all()
        
        # マッチング処理
        matches = []
        
        for applicant in applicants:
            # 求職者の年齢を計算
            applicant_age = None
            if applicant.birthdate:
                applicant_age = calculate_age(applicant.birthdate)
            
            match_data = {
                'applicant': applicant.to_dict(),
                'age_match': True,
                'location_match': 0,
                'age': applicant_age
            }
            
            # 1. 年齢条件の確認
            if applicant_age is not None and job.age_limit is not None:
                if applicant_age > job.age_limit:
                    match_data['age_match'] = False
            
            # 2. 希望勤務地の一致確認
            if hasattr(applicant, 'desired_location') and applicant.desired_location and job.prefecture:
                # 希望勤務地と都道府県の一致度を評価
                if applicant.desired_location in job.prefecture or job.prefecture in applicant.desired_location:
                    match_data['location_match'] = 100  # 完全一致
                # 部分一致の確認
                elif any(loc in job.prefecture for loc in extract_keywords(applicant.desired_location)) or \
                     any(loc in applicant.desired_location for loc in extract_keywords(job.prefecture)):
                    match_data['location_match'] = 50  # 部分一致
            
            # マッチリストに追加
            matches.append(match_data)
        
        # ソート処理
        if sort_by == 'location':
            # 希望勤務地マッチ度でソート（高いものが先）
            matches = sorted(matches, key=lambda x: (x['location_match']), reverse=True)
        else:
            # デフォルトは年齢順（若い順）、年齢が不明のものは後ろに
            matches = sorted(matches, key=lambda x: (
                0 if x['age'] is None else 1,  # 年齢がある人が先
                x['age'] if x['age'] is not None else 999  # 若い順
            ))
        
        app.logger.info(f'求人マッチング結果: ID={job_id}, 件数={len(matches)}')
        response = make_response(jsonify(matches))
        return add_cors_headers(response)
    except Exception as e:
        app.logger.exception(f"求人マッチングエラー: ID={job_id}, エラー={str(e)}")
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)

if __name__ == '__main__':
    if DEBUG:
        app.run(debug=True, host='0.0.0.0', port=5001)
    else:
        # 本番環境では、gunicornなどのWSGIサーバーを使用することを推奨
        app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5001)))

# Google Distance Matrix APIを使用した距離計算エンドポイント
@app.route('/api/calculate-distances', methods=['POST', 'OPTIONS'])
def calculate_distances():
    # OPTIONSリクエストを明示的に処理
    if request.method == 'OPTIONS':
        app.logger.info(f'OPTIONS リクエスト処理: /api/calculate-distances')
        response = make_response()
        return add_cors_headers(response)
    
    try:
        data = request.json
        app.logger.info(f"距離計算リクエスト受信: {data}")
        
        # 以下は既存のコード
        origin = data.get('origin')
        destinations = data.get('destinations')
        mode = data.get('mode', 'driving')  # デフォルトは車
        
        if not origin or not destinations:
            response = make_response(jsonify({'error': '出発地と目的地の指定が必要です'}), 400)
            return add_cors_headers(response)
        
        # Google Distance Matrix API接続用のパラメータ
        api_key = os.environ.get('GOOGLE_MAPS_API_KEY')
        if not api_key:
            response = make_response(jsonify({'error': 'Google Maps APIキーが設定されていません'}), 500)
            return add_cors_headers(response)
            
        # APIリクエストの準備
        import googlemaps
        
        gmaps = googlemaps.Client(key=api_key)
        
        # 目的地は複数あるためバッチで処理
        batch_size = 25  # Google APIの制限に合わせて調整
        batches = [destinations[i:i + batch_size] for i in range(0, len(destinations), batch_size)]
        
        all_distances = []
        all_durations = []
        all_distance_texts = []
        all_duration_texts = []
        
        for batch in batches:
            try:
                # Google APIを呼び出し
                matrix = gmaps.distance_matrix(
                    origin,
                    batch,
                    mode=mode,
                    language="ja",
                    units="metric"
                )
                
                # 結果を解析
                if matrix['status'] == 'OK':
                    for element in matrix['rows'][0]['elements']:
                        if element['status'] == 'OK':
                            all_distances.append(element['distance']['value'])  # メートル単位
                            all_durations.append(element['duration']['value'])  # 秒単位
                            all_distance_texts.append(element['distance']['text'])  # 表示用テキスト
                            all_duration_texts.append(element['duration']['text'])  # 表示用テキスト
                        else:
                            all_distances.append(None)
                            all_durations.append(None)
                            all_distance_texts.append('計算不能')
                            all_duration_texts.append('計算不能')
                else:
                    raise Exception(f"Google API error: {matrix['status']}")
                
            except Exception as batch_error:
                app.logger.error(f"バッチ処理エラー: {str(batch_error)}")
                # エラーが発生した場合はダミーデータで埋める
                all_distances.extend([None] * len(batch))
                all_durations.extend([None] * len(batch))
                all_distance_texts.extend(['計算不能'] * len(batch))
                all_duration_texts.extend(['計算不能'] * len(batch))
        
        response = make_response(jsonify({
            'distances': all_distances,
            'durations': all_durations,
            'distanceTexts': all_distance_texts,
            'durationTexts': all_duration_texts
        }))
        return add_cors_headers(response)
        
    except Exception as e:
        app.logger.exception(f"距離計算エラー: {str(e)}\n{traceback.format_exc()}")
        response = make_response(jsonify({'error': str(e)}), 500)
        return add_cors_headers(response)
    


# アプリケーションコンテキスト内でデータベーステーブルを更新するためのルート
@app.route('/api/update-database', methods=['POST'])
def update_database():
    try:
        # 既存のテーブルを保持したままカラムを追加
        with app.app_context():
            # 前の状態を保存
            inspector = db.inspect(db.engine)
            existing_tables = inspector.get_table_names()
            existing_columns = {}
            
            for table in existing_tables:
                existing_columns[table] = [col['name'] for col in inspector.get_columns(table)]
            
            # モデルの更新を反映
            db.reflect()
            db.create_all()
            
            # 新しいカラムが追加されたかログ出力
            for table in existing_tables:
                new_columns = [col['name'] for col in inspector.get_columns(table)]
                added_columns = set(new_columns) - set(existing_columns[table])
                if added_columns:
                    app.logger.info(f"テーブル {table} に新しいカラムが追加されました: {added_columns}")
        
        return jsonify({"message": "データベースが正常に更新されました"}), 200
    except Exception as e:
        app.logger.error(f"データベース更新エラー: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/debug/applicant-model', methods=['GET'])
def debug_applicant_model():
    """Applicantモデルとデータベース構造を診断するためのデバッグエンドポイント"""
    try:
        import inspect
        import sqlalchemy as sa
        
        debug_info = {
            "database_info": {},
            "model_info": {},
            "sample_data": {},
            "error_diagnosis": []
        }
        
        # データベース構造の確認
        inspector = db.inspect(db.engine)
        tables = inspector.get_table_names()
        debug_info["database_info"]["tables"] = tables
        
        if "applicant" in tables:
            columns = inspector.get_columns("applicant")
            debug_info["database_info"]["applicant_columns"] = [
                {"name": col["name"], "type": str(col["type"])} 
                for col in columns
            ]
            
            # 'assigned_employee_id'カラムの存在確認
            assigned_employee_column = next(
                (col for col in columns if col["name"] == "assigned_employee_id"), 
                None
            )
            if assigned_employee_column:
                debug_info["database_info"]["has_assigned_employee_id"] = True
            else:
                debug_info["database_info"]["has_assigned_employee_id"] = False
                debug_info["error_diagnosis"].append(
                    "データベースに'assigned_employee_id'カラムが存在しません。"
                )
        else:
            debug_info["error_diagnosis"].append("'applicant'テーブルが見つかりません。")
        
        # Applicantモデルの確認
        try:
            model_attrs = {
                name: attr for name, attr in inspect.getmembers(Applicant)
                if not name.startswith('_') and not inspect.ismethod(attr)
            }
            
            debug_info["model_info"]["attributes"] = list(model_attrs.keys())
            
            # __table__情報の取得
            if hasattr(Applicant, '__table__'):
                table_columns = [c.name for c in Applicant.__table__.columns]
                debug_info["model_info"]["table_columns"] = table_columns
                
                if "assigned_employee_id" in table_columns:
                    debug_info["model_info"]["has_assigned_employee_id_in_model"] = True
                else:
                    debug_info["model_info"]["has_assigned_employee_id_in_model"] = False
                    debug_info["error_diagnosis"].append(
                        "Applicantモデルに'assigned_employee_id'フィールドが定義されていません。"
                    )
        except Exception as model_err:
            debug_info["model_info"]["error"] = str(model_err)
            debug_info["error_diagnosis"].append(
                f"モデル情報取得エラー: {str(model_err)}"
            )
        
        # サンプルデータの取得
        try:
            applicant = Applicant.query.first()
            if applicant:
                applicant_dict = {
                    c.name: getattr(applicant, c.name)
                    for c in Applicant.__table__.columns
                }
                debug_info["sample_data"]["first_applicant"] = applicant_dict
            else:
                debug_info["sample_data"]["first_applicant"] = None
                debug_info["error_diagnosis"].append("Applicantデータが存在しません。")
        except Exception as data_err:
            debug_info["sample_data"]["error"] = str(data_err)
            debug_info["error_diagnosis"].append(
                f"サンプルデータ取得エラー: {str(data_err)}"
            )
        
        # 問題の診断と解決策
        if not debug_info["error_diagnosis"]:
            debug_info["status"] = "OK"
            debug_info["message"] = "Applicantモデルとデータベース構造に問題はありません。"
        else:
            debug_info["status"] = "ERROR"
            debug_info["message"] = "いくつかの問題が見つかりました。"
            
            # 解決策の提案
            debug_info["solutions"] = []
            
            if not debug_info["database_info"].get("has_assigned_employee_id", False):
                debug_info["solutions"].append(
                    "データベースに'assigned_employee_id'カラムを追加してください: "
                    "ALTER TABLE applicant ADD COLUMN assigned_employee_id INTEGER;"
                )
            
            if not debug_info["model_info"].get("has_assigned_employee_id_in_model", False):
                debug_info["solutions"].append(
                    "models.pyのApplicantクラスに'assigned_employee_id'フィールドを追加してください: "
                    "assigned_employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=True)"
                )
            
            debug_info["solutions"].append(
                "アプリケーションサーバーを再起動して変更を反映させてください。"
            )
        
        return jsonify(debug_info), 200
    
    except Exception as e:
        import traceback
        return jsonify({
            "status": "ERROR",
            "message": f"デバッグ中にエラーが発生しました: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500
